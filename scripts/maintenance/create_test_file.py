#!/usr/bin/env python3
"""
Generate test Excel file for PO upload testing
"""

import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[ERROR] openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

def create_test_po_file(filename="test_po.xlsx"):
    """Create a test PO Excel file"""
    
    print(f"Creating test PO file: {filename}")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Data"
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define headers
    headers = [
        "PO Number",
        "Vendor Name",
        "Date",
        "Amount",
        "Currency",
        "Status",
        "Items",
        "Description",
        "Location",
        "Project"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add sample data rows
    sample_data = [
        ("PO-2024-001", "Vendor A", "2024-01-15", 50000, "USD", "Open", "5", "Raw Materials", "Mumbai", "Project Alpha"),
        ("PO-2024-002", "Vendor B", "2024-01-16", 75000, "USD", "Partial", "3", "Spare Parts", "Delhi", "Project Beta"),
        ("PO-2024-003", "Vendor C", "2024-01-17", 120000, "USD", "Closed", "8", "Equipment", "Bangalore", "Project Gamma"),
        ("PO-2024-004", "Vendor D", "2024-01-18", 45000, "USD", "Open", "2", "Services", "Chennai", "Project Delta"),
        ("PO-2024-005", "Vendor E", "2024-01-19", 95000, "USD", "Pending", "10", "Consumables", "Pune", "Project Epsilon"),
        ("PO-2024-006", "Vendor A", "2024-01-20", 67500, "USD", "Open", "4", "Raw Materials", "Mumbai", "Project Zeta"),
        ("PO-2024-007", "Vendor F", "2024-01-21", 150000, "USD", "Processing", "15", "Heavy Equipment", "Kolkata", "Project Eta"),
        ("PO-2024-008", "Vendor G", "2024-01-22", 32000, "USD", "Open", "2", "Tools", "Hyderabad", "Project Theta"),
        ("PO-2024-009", "Vendor B", "2024-01-23", 88000, "USD", "Partial", "6", "Components", "Ahmedabad", "Project Iota"),
        ("PO-2024-010", "Vendor H", "2024-01-24", 145000, "USD", "Closed", "12", "Assembly", "Surat", "Project Kappa"),
    ]
    
    # Write data rows
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            if col_num == 4:  # Amount column
                cell.number_format = '#,##0.00'
            elif col_num == 3:  # Date column
                cell.number_format = 'YYYY-MM-DD'
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Adjust column widths
    column_widths = [15, 15, 12, 12, 10, 12, 8, 20, 12, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save file
    wb.save(filename)
    print(f"[OK] Test file created: {filename}")
    print(f"[INFO] File contains {len(sample_data)} sample PO records")
    return filename

if __name__ == "__main__":
    try:
        create_test_po_file()
        print("\n[SUCCESS] Test file ready for upload testing!")
        print("You can now run: python test_all_routes.py")
    except Exception as e:
        print(f"[ERROR] Failed to create test file: {e}")
        sys.exit(1)
