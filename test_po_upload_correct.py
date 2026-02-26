#!/usr/bin/env python3
"""
Test script showing the correct way to send files to /api/po/upload endpoint

The endpoint expects:
- file: In multipart/form-data body (field name: "file")
- client_id: As URL query parameter (?client_id=1)
- Other metadata: As URL query parameters
"""

import requests
import sys

def test_po_upload(base_url="http://localhost:8000"):
    """Test PO upload with proper multipart/form-data and query parameters"""
    
    url = f"{base_url}/api/po/upload"
    
    # Prepare query parameters
    params = {
        "client_id": 1,  # Bajaj
        "auto_save": True
    }
    
    # Create a test file in memory
    test_file_content = b"%PDF-1.4\n%Test PDF content"
    test_filename = "test_bajaj_po.pdf"
    
    print(f"🧪 Testing file upload to {url}")
    print(f"   Query params: {params}")
    print(f"   File name: {test_filename}")
    print(f"   File size: {len(test_file_content)} bytes\n")
    
    try:
        # IMPORTANT: Use 'files' parameter with field name 'file'
        files = {
            'file': (test_filename, test_file_content, 'application/pdf')
        }
        
        # Send POST request with files and query parameters
        response = requests.post(url, params=params, files=files, timeout=30)
        
        print(f"Status Code: {response.status_code}\n")
        
        if response.status_code == 200:
            result = response.json()
            print("✅ SUCCESS!\n")
            print(f"Session ID: {result.get('session_id')}")
            print(f"Client ID: {result.get('client_id')}")
            print(f"File ID: {result.get('file_id')}")
            print(f"Parser Type: {result.get('parser_type')}")
            return True
        else:
            print("❌ ERROR\n")
            print("Response:")
            print(response.json())
            return False
            
    except Exception as e:
        print(f"❌ Request failed: {e}")
        return False


def test_with_excel():
    """Test with an Excel file (requires openpyxl)"""
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl not installed, skipping Excel test")
        return
    
    from io import BytesIO
    
    print("\n" + "="*60)
    print("Testing with Excel file")
    print("="*60 + "\n")
    
    url = "http://localhost:8000/api/po/upload"
    params = {
        "client_id": 1,
        "project_name": "Test Project",
        "auto_save": True
    }
    
    # Create a minimal Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "PO Number"
    ws['B1'] = "Amount"
    ws['A2'] = "PO-12345"
    ws['B2'] = 1000
    
    excel_content = BytesIO()
    wb.save(excel_content)
    excel_content.seek(0)
    
    print(f"Testing file upload to {url}")
    print(f"Query params: {params}\n")
    
    try:
        files = {
            'file': ('test_po.xlsx', excel_content.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        response = requests.post(url, params=params, files=files, timeout=30)
        
        print(f"Status Code: {response.status_code}\n")
        
        if response.status_code == 200:
            result = response.json()
            print("✅ SUCCESS!\n")
            print(f"File ID: {result.get('file_id')}")
            print(f"Line Items Count: {result.get('line_item_count')}")
            print(f"PO Details: {result.get('po_details')}")
        else:
            print("❌ ERROR")
            print(response.json())
            
    except Exception as e:
        print(f"❌ Request failed: {e}")


if __name__ == "__main__":
    # Run basic test
    success = test_po_upload()
    
    # Run Excel test
    test_with_excel()
    
    sys.exit(0 if success else 1)
