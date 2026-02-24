from openpyxl import load_workbook
from datetime import datetime
import re
import json


# -------------------------------------------------------
# Custom Exception
# -------------------------------------------------------

class BajajPOParserError(Exception):
    """Exception raised for Bajaj PO parsing errors"""
    pass


# -------------------------------------------------------
# Helpers
# -------------------------------------------------------

def clean(v):
    if v is None:
        return ""
    return " ".join(str(v).replace("\n", " ").strip().split())


def is_number(v):
    if v is None:
        return False
    try:
        s = str(v)
        s = re.sub(r"[,\s₹$£€]", "", s)
        s = s.replace("(", "-").replace(")", "")
        float(s)
        return True
    except:
        return False


def to_float(v):
    if not is_number(v):
        return None
    s = str(v)
    s = re.sub(r"[,\s₹$£€]", "", s)
    s = s.replace("(", "-").replace(")", "")
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s)
    except:
        return None


# -------------------------------------------------------
# Safe Site Address Extractor
# -------------------------------------------------------

def extract_site_address(ws):

    keywords = ("SHIP TO", "DELIVERY TO", "SITE ADDRESS", "DELIVERY ADDRESS")

    for r in range(1, min(12, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):

            raw = clean(ws.cell(r, c).value)
            if not raw:
                continue

            upper = raw.upper()

            # Avoid table headers
            if "DESCRIPTION" in upper:
                continue

            if any(k in upper for k in keywords):

                # right cell
                right = clean(ws.cell(r, c + 1).value)
                if right and "DESCRIPTION" not in right.upper():
                    return right

                # below cell
                below = clean(ws.cell(r + 1, c).value)
                if below and "DESCRIPTION" not in below.upper():
                    return below

                # inline after colon
                parts = re.split(r"[:\-]\s*", raw, maxsplit=1)
                if len(parts) == 2 and "DESCRIPTION" not in parts[1].upper():
                    return parts[1].strip()

    return None


# -------------------------------------------------------
# Store ID extractor
# -------------------------------------------------------

def extract_store_id(ws, address_hint=None):

    if address_hint:
        m = re.search(r"\b[A-Z]{1,4}\d{1,6}\b", address_hint.upper())
        if m:
            return m.group(0)

    text = " ".join(
        clean(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 80) + 1)
        for c in range(1, ws.max_column + 1)
    ).upper()

    m = re.search(r"(STORE|OUTLET|BRANCH)\s*[:\-]?\s*([A-Z0-9\-]{3,20})", text)
    if m:
        return m.group(2)

    tokens = re.findall(r"\b[A-Z]{1,4}\d{1,6}\b", text)
    if tokens:
        return sorted(tokens, key=len, reverse=True)[0]

    return None


# -------------------------------------------------------
# Main Reliable Parser
# -------------------------------------------------------

def parse_bajaj_po(file_path):

    warnings = []

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        return {
            "status": "FILE_READ_ERROR",
            "error": str(e),
            "po_details": {},
            "line_items": [],
            "warnings": ["Workbook could not be opened"]
        }

    # -------------------------
    # Address & Store
    # -------------------------

    site_address = extract_site_address(ws)

    if not site_address:
        site_address = "UNKNOWN_SITE"
        warnings.append("Site address not detected → fallback used")

    store_id = extract_store_id(ws, site_address)

    if not store_id:
        store_id = "UNKNOWN_STORE"
        warnings.append("Store ID not detected → fallback used")

    # -------------------------
    # PO number & date
    # -------------------------

    po_number = None
    po_date = None

    for r in range(1, min(ws.max_row, 50) + 1):
        for c in range(1, ws.max_column + 1):

            v = ws.cell(r, c).value
            if not v:
                continue

            s = clean(v)

            if not po_number:
                m = re.search(r"\b\d{6,12}\b", s.replace(" ", ""))
                if m:
                    po_number = m.group(0)

            if not po_date and isinstance(v, datetime):
                po_date = v.date().isoformat()

        if po_number and po_date:
            break

    if not po_number:
        po_number = "UNKNOWN_PO"
        warnings.append("PO number not detected")

    # -------------------------
    # Header detection
    # -------------------------

    header_row = None
    desc_col = None
    qty_col = None

    for r in range(1, min(ws.max_row, 80) + 1):

        row_vals = [clean(ws.cell(r, c).value).upper()
                    for c in range(1, ws.max_column + 1)]

        if "DESCRIPTION" in " ".join(row_vals):

            header_row = r

            for c, val in enumerate(row_vals, start=1):
                if "DESCRIPTION" in val:
                    desc_col = c
                if val in ("QTY", "QUANTITY"):
                    qty_col = c

            break

    if header_row is None:
        warnings.append("Line item header not detected")
        return {
            "po_details": {
                "po_number": po_number,
                "po_date": po_date,
                "store_id": store_id,
                "site_address": site_address
            },
            "summary": {"grand_total": None},
            "line_items": [],
            "line_item_count": 0,
            "warnings": warnings
        }

    start_row = header_row + 1

    # -------------------------
    # Amount column detection
    # -------------------------

    col_scores = []

    for c in range(desc_col + 1, ws.max_column + 1):

        cnt = 0
        total = 0

        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(r, c).value
            if is_number(v):
                cnt += 1
                total += to_float(v) or 0

        col_scores.append((c, cnt, total))

    col_scores.sort(key=lambda x: (x[1], x[2]), reverse=True)

    amount_col = col_scores[0][0] if col_scores else None

    if not amount_col:
        warnings.append("Amount column not detected")

    # -------------------------
    # Parse items
    # -------------------------

    items = []

    for r in range(start_row, ws.max_row + 1):

        desc = clean(ws.cell(r, desc_col).value)
        if not desc:
            continue

        if desc.upper() == "TOTAL":
            continue

        amt = to_float(ws.cell(r, amount_col).value) if amount_col else None
        if amt is None:
            continue

        qty = 1
        if qty_col:
            q = to_float(ws.cell(r, qty_col).value)
            if q:
                qty = q

        items.append({
            "description": desc,
            "quantity": qty,
            "amount": round(amt, 2),
            "row_index": r
        })

    grand_total = round(sum(i["amount"] for i in items), 2) if items else None

    if not items:
        warnings.append("No line items parsed")

    # -------------------------
    # Final JSON
    # -------------------------

    return {
        "status": "PARSE_SUCCESS",
        "po_details": {
            "po_number": po_number,
            "po_date": po_date,
            "store_id": store_id,
            "site_address": site_address
        },
        "summary": {
            "grand_total": grand_total
        },
        "line_items": items,
        "line_item_count": len(items),
        "warnings": warnings
    }


# -------------------------------------------------------
# Example run
# -------------------------------------------------------

if __name__ == "__main__":
    path = "/mnt/data/BAJAJ2.xlsx"
    result = parse_bajaj_po(path)
    print(json.dumps(result, indent=2))
