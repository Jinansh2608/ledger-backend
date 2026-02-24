from openpyxl import load_workbook
import re
from typing import Dict, Any, List, Tuple


class ProformaInvoiceParserError(Exception):
    pass


# -----------------------------------------------------
# Helpers
# -----------------------------------------------------

def clean(v):
    if v is None:
        return ""
    return " ".join(str(v).replace("\n", " ").split()).strip()


def is_number(v):
    try:
        if v is None or str(v).strip() == "":
            return False
        # allow numbers with commas, parentheses for negatives, and decimals
        s = str(v).replace(",", "").replace("(", "-").replace(")", "")
        float(s)
        return True
    except:
        return False


def to_float(v):
    try:
        if v is None or str(v).strip() == "":
            return 0.0
        s = str(v).replace(",", "").replace("(", "-").replace(")", "")
        return float(s)
    except:
        return 0.0


# Robust cell getter to handle merged-like blanks (look upward)
def get_cell(ws, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    if v not in (None, ""):
        return clean(v)
    # look upward a few rows (merged cells or messy docs)
    rr = r - 1
    while rr >= 1 and r - rr <= 5:  # limit the upward search to 5 rows
        vv = ws.cell(rr, c).value
        if vv not in (None, ""):
            return clean(vv)
        rr -= 1
    return ""


# -----------------------------------------------------
# Header detection
# -----------------------------------------------------

def find_header_row(ws) -> int:
    """
    Find BOQ header row using semantic markers.
    Uses get_cell to be robust to merged/blank cells.
    """
    for r in range(1, ws.max_row + 1):
        row_vals = [get_cell(ws, r, c) for c in range(1, ws.max_column + 1)]
        row = " ".join(row_vals).upper()
        if (
            (re.search(r"\b(SR|SL|NO\.?|1\.)\b", row) or re.search(r"^\s*1\s+", row)) and
            (re.search(r"\b(BOQ|DESC|PARTICULARS|ITEM|PRODUCT|DESCRIPTION)\b", row)) and
            (re.search(r"\b(QTY|QUAN|QTY\.)\b", row) or re.search(r"\b(RATE|PRICE|UNIT PRICE)\b", row))
        ):
            return r
    return None


def extract_store_id(ws) -> str:
    text = " ".join(
        get_cell(ws, r, c)
        for r in range(1, 40)
        for c in range(1, ws.max_column + 1)
    )

    patterns = [
        r"(?:STORE|SITE|LOCATION|OUTLET|BRANCH|UNIT)\s*(?:ID|CODE|NO|#|REF)?\s*[:\-]?\s*([A-Z0-9\/\-\_]+)",
        r"(?:STORE|SITE)\s*[:\-]\s*([A-Z0-9\/\-\_]+)",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            return m.group(1)
    return ""


def parse_date(date_str):
    """Parse date in DD-MM-YYYY or DD/MM/YYYY format to YYYY-MM-DD format. Returns None if invalid."""
    if not date_str:
        return None
    try:
        date_str = str(date_str).strip()
        if date_str.count('-') == 2 or date_str.count('/') == 2:
            sep = '-' if '-' in date_str else '/'
            parts = date_str.split(sep)
            if len(parts) == 3:
                day, month, year = parts
                day_int = int(re.sub(r"\D", "", day) or 0)
                month_int = int(re.sub(r"\D", "", month) or 0)
                year_int = int(re.sub(r"\D", "", year) or 0)
                if 1 <= day_int <= 31 and 1 <= month_int <= 12:
                    return f"{year_int:04d}-{month_int:02d}-{day_int:02d}"
    except:
        pass
    return None


def extract_header_fields(ws) -> Dict[str, Any]:
    header = {
        "vendor_name": "",
        "vendor_address": "",
        "vendor_gstin": "",
        "client_name": "",
        "client_po_number": "",
        "po_number": "",
        "po_date": None,
        "pi_number": "",
        "pi_date": None,
        "bill_to_gstin": "",
        "bill_to_address": "",
        "ship_to_address": "",
        "site_name": "",
    }

    # scan rows and nearby columns; allow value in col 2..4
    for r in range(1, min(31, ws.max_row + 1)):
        label = get_cell(ws, r, 1).upper()
        # try also columns 1..3 as label/value pairs
        values = [get_cell(ws, r, c) for c in range(2, min(5, ws.max_column + 1))]
        value = next((v for v in values if v), "")

        if not label and not value:
            continue

        # vendor
        if ("VENDOR" in label or "SUPPLIER" in label) and any(t in label for t in ("CO.", "NAME", "COMP")):
            header["vendor_name"] = value
        elif ("VENDOR" in label or "SUPPLIER" in label) and any(t in label for t in ("ADDRESS", "ADD", "ADDR")):
            header["vendor_address"] = value
        elif ("VENDOR" in label or "SUPPLIER" in label) and any(t in label for t in ("GSTIN", "GST", "TAX")):
            header["vendor_gstin"] = value

        # client/bill
        elif any(t in label for t in ("BILL", "CLIENT")) and any(s in label for s in ("TO", "NAME")):
            header["client_name"] = value
        elif "BILL" in label and any(t in label for t in ("ADDRESS", "ADD", "ADDR")):
            header["bill_to_address"] = value
        elif "BILL" in label and any(t in label for t in ("GSTIN", "GST", "TAX")):
            header["bill_to_gstin"] = value

        # ship to
        elif "SHIP" in label and any(t in label for t in ("TO", "NAME", "ADDRESS", "ADD", "ADDR")):
            header["ship_to_address"] = value

        # PO / ref
        elif any(x in label for x in ("PO", "PURCHASE", "ORDER", "REF")) and any(x in label for x in ("NO", "NUMBER", "REF", "#")) and "DATE" not in label:
            header["client_po_number"] = value
            header["po_number"] = value
        elif "REF" in label and "NO" in label:
            header["client_po_number"] = value
            header["po_number"] = value

        # PI / invoice
        elif any(x in label for x in ("PI", "INVOICE", "PROFORMA")) and any(x in label for x in ("NO", "NUMBER", "#")) and "DATE" not in label:
            header["pi_number"] = value
        elif any(x in label for x in ("PI", "INVOICE", "PROFORMA")) and "DATE" in label:
            header["pi_date"] = parse_date(value)

        # site
        elif any(x in label for x in ("SITE", "LOCATION")) and any(x in label for x in ("NAME", "ID")) and "ID" not in label:
            header["site_name"] = value

    # fallback: scan larger header text for PO/Ref
    if not header["po_number"]:
        header_text = " ".join(
            get_cell(ws, r, c)
            for r in range(1, 40)
            for c in range(1, ws.max_column + 1)
        )
        po_patterns = [
            r"(?:PO|P\.O\.?|PURCHASE\s*ORDER)\s*(?:NO|NUMBER|#|REF)?\s*[:\-]?\s*([A-Z0-9\/\-\_]+)",
            r"(?:CLIENT\s*REF|CUSTOMER\s*REF|REF|REFERENCE)\s*[:\-]?\s*([A-Z0-9\/\-\_]+)",
        ]
        for pat in po_patterns:
            m = re.search(pat, header_text, re.I)
            if m:
                header["po_number"] = m.group(1)
                header["client_po_number"] = m.group(1)
                break

    if header["client_po_number"] and not header["po_number"]:
        header["po_number"] = header["client_po_number"]

    return header


# -----------------------------------------------------
# Column mapping (with inference)
# -----------------------------------------------------

def map_columns(ws, header_row: int) -> Dict[str, int]:
    headers = {
        c: get_cell(ws, header_row, c).upper()
        for c in range(1, ws.max_column + 1)
    }
    cols: Dict[str, int] = {}

    for c, txt in headers.items():
        if any(x in txt for x in ["SR", "SL", "NO."]) and "GST" not in txt:
            cols["sr"] = c
        elif any(x in txt for x in ["BOQ", "DESC", "NAME", "PARTICULARS", "ITEM", "PRODUCT", "DESCRIPTION"]):
            cols["desc"] = c
        elif "HSN" in txt:
            cols["hsn"] = c
        elif any(x in txt for x in ["QTY", "QUAN"]):
            cols["qty"] = c
        elif "UNIT" in txt and "PRICE" not in txt:
            cols["unit"] = c
        elif any(x in txt for x in ["RATE", "PRICE", "UNIT PRICE"]) and "TOTAL" not in txt:
            cols["rate"] = c
        elif "TOTAL WITH" in txt or ("TOTAL" in txt and "GST" in txt):
            cols["total_with_gst"] = c
        elif any(x in txt for x in ["TOTAL", "AMOUNT", "VALUE", "NET"]) and "WITH" not in txt and "GST" not in txt and "TAX" not in txt:
            cols["total"] = c
        elif any(x in txt for x in ["GST", "TAX"]) and "TOTAL" not in txt and "RATE" not in txt:
            cols["tax_amount"] = c

    # If core columns missing, try to infer numeric-heavy columns by scanning next few rows
    needed = ["qty", "rate", "total"]
    missing = [k for k in needed if k not in cols]
    if missing:
        numeric_counts = {c: 0 for c in range(1, ws.max_column + 1)}
        sample_rows = range(header_row + 1, min(header_row + 20, ws.max_row + 1))
        for r in sample_rows:
            for c in range(1, ws.max_column + 1):
                v = get_cell(ws, r, c)
                if is_number(v):
                    numeric_counts[c] += 1
        # pick columns with highest numeric counts for missing fields (simple heuristic)
        sorted_cols = sorted(numeric_counts.items(), key=lambda x: x[1], reverse=True)
        for k in missing:
            for c, count in sorted_cols:
                if count <= 0:
                    break
                if c not in cols.values():
                    cols[k] = c
                    break

    return cols


# -----------------------------------------------------
# Item extraction (robust)
# -----------------------------------------------------

def looks_like_total_row(row_vals: List[str]) -> bool:
    joined = " ".join(row_vals).upper()
    if re.search(r"\b(TOTAL|GRAND TOTAL|NET PAYABLE|SUBTOTAL|AMOUNT DUE)\b", joined):
        # ensure there is a numeric value in the row
        if any(is_number(x) for x in row_vals):
            return True
    return False


def extract_items(ws, start_row: int, cols: Dict[str, int]) -> Tuple[List[Dict[str, Any]], int]:
    items: List[Dict[str, Any]] = []
    current: Dict[str, Any] = None
    r = start_row

    # safety: if sr/desc missing, attempt to find a desc column via first non-empty column in header_row region
    while r <= ws.max_row:
        row_vals = [get_cell(ws, r, c) for c in range(1, ws.max_column + 1)]
        row_text = " ".join(row_vals).upper()

        # stop when we hit summary/total row
        if looks_like_total_row(row_vals):
            break

        # find candidate columns values using cols mapping (fallback to empty if not mapped)
        sr = row_vals[cols["sr"] - 1] if "sr" in cols else ""
        desc = row_vals[cols["desc"] - 1] if "desc" in cols else ""
        qty = row_vals[cols["qty"] - 1] if "qty" in cols else ""
        rate = row_vals[cols["rate"] - 1] if "rate" in cols else ""
        total = row_vals[cols["total"] - 1] if "total" in cols else ""
        tax_amount = row_vals[cols.get("tax_amount", 1) - 1] if "tax_amount" in cols else ""
        total_with_gst = row_vals[cols.get("total_with_gst", 1) - 1] if "total_with_gst" in cols else ""

        begins_item = False

        # rule 1: SR present and numeric -> new item
        if is_number(sr):
            begins_item = True

        # rule 2: if SR absent but desc present and (qty or rate or total numeric) -> new item
        elif desc and (is_number(qty) or is_number(rate) or is_number(total) or is_number(total_with_gst)):
            begins_item = True

        # rule 3: sometimes row starts with a numbered bullet "1." or "1)" in first cell
        elif re.match(r"^\d+[\.\)]", row_vals[0]):
            begins_item = True

        if begins_item:
            # push previous
            if current:
                items.append(current)

            gross = to_float(total_with_gst) if total_with_gst else (to_float(total) + to_float(tax_amount))
            current = {
                "sr": int(float(sr)) if is_number(sr) else None,
                "boq_name": desc or "",
                "hsn_code": row_vals[cols.get("hsn", 1) - 1] if "hsn" in cols else "",
                "quantity": to_float(qty),
                "unit": row_vals[cols.get("unit", 1) - 1] if "unit" in cols else "",
                "rate": to_float(rate),
                "taxable_amount": to_float(total),
                "tax_rate": row_vals.get(cols.get("tax_rate", None) - 1, "") if "tax_rate" in cols else "",
                "tax_amount": to_float(tax_amount),
                "gst_amount": to_float(tax_amount),
                "total_with_gst": to_float(total_with_gst),
                "gross_amount": gross,
            }

        else:
            # continuation line: append textual pieces to last item's description
            if current:
                # try to find a textual fragment in the row: choose first non-numeric cell which is likely a description continuation
                continuation = ""
                for cell in row_vals:
                    if cell and not is_number(cell) and not re.search(r"\b(QTY|RATE|AMOUNT|TOTAL|GST|CGST|SGST|IGST)\b", cell.upper()):
                        continuation = cell
                        break
                if continuation:
                    current["boq_name"] = (current.get("boq_name", "") + " " + continuation).strip()
            # else: stray row before any item - skip

        r += 1

    # append last
    if current:
        items.append(current)

    return items, r


# -----------------------------------------------------
# Summary extraction
# -----------------------------------------------------

def extract_summary(ws, start_row: int) -> Dict[str, float]:
    summary = {
        "subtotal": 0.0,
        "cgst": 0.0,
        "sgst": 0.0,
        "igst": 0.0,
        "total_amount": 0.0,
    }

    for r in range(start_row, ws.max_row + 1):
        row = [get_cell(ws, r, c) for c in range(1, ws.max_column + 1)]
        text = " ".join(row).upper()
        nums = [to_float(v) for v in row if is_number(v)]
        if not nums:
            continue
        val = nums[-1]
        if re.search(r"^\s*TOTAL\b", text):
            summary["subtotal"] = val
        elif "CGST" in text:
            summary["cgst"] = val
        elif "SGST" in text:
            summary["sgst"] = val
        elif "IGST" in text:
            summary["igst"] = val
        elif re.search(r"(PI TOTAL|GRAND TOTAL|NET PAYABLE|AMOUNT DUE)", text):
            summary["total_amount"] = val

    return summary


# -----------------------------------------------------
# Main parser
# -----------------------------------------------------

def parse_proforma_invoice(path: str, debug: bool = False) -> Dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)
    if not header_row:
        raise ProformaInvoiceParserError("BOQ header not found")

    header_fields = extract_header_fields(ws)

    store_id = extract_store_id(ws)
    if store_id:
        header_fields["store_id"] = store_id

    cols = map_columns(ws, header_row)
    if debug:
        print("Detected columns:", cols)

    items, summary_start = extract_items(ws, header_row + 1, cols)
    if debug:
        print(f"Parsed {len(items)} line items. Summary starts at row {summary_start}")

    summary = extract_summary(ws, summary_start)

    po_details = {
        **header_fields,
        **summary
    }

    return {
        "po_details": po_details,
        "line_items": items,
        "line_item_count": len(items)
    }
