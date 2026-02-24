#!/usr/bin/env python3
"""
Test file upload system - comprehensive test script
Creates sample data and tests all file upload routes
"""
import requests
import json
import time
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

BASE_URL = "http://127.0.0.1:8000/api"

def print_section(title):
    """Print a formatted section header"""
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}\n")

def create_sample_bajaj_po():
    """Create a sample Bajaj PO Excel file"""
    print("Creating sample Bajaj PO Excel file...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO"
    
    # Headers
    headers = ["PO Number", "PO Date", "Vendor", "Item Code", "Description", "Quantity", "Unit Price", "Total"]
    ws.append(headers)
    
    # Format header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data
    po_data = [
        ["PO-2026-001", "2026-02-17", "Vendor ABC", "ITEM001", "Widget A", 100, 50.00, 5000.00],
        ["PO-2026-001", "2026-02-17", "Vendor ABC", "ITEM002", "Widget B", 50, 75.00, 3750.00],
        ["PO-2026-001", "2026-02-17", "Vendor ABC", "ITEM003", "Widget C", 25, 100.00, 2500.00],
    ]
    
    for row in po_data:
        ws.append(row)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Save to temp file
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    print(f"✓ Created sample PO file: {temp_file.name}")
    return temp_file.name

def test_health_check():
    """Test health endpoint"""
    print_section("1. HEALTH CHECK")
    try:
        resp = requests.get(f"{BASE_URL}/health")
        print(f"Status: {resp.status_code}")
        print(f"Response: {json.dumps(resp.json(), indent=2)}")
        return resp.status_code == 200
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_get_clients():
    """Test get supported clients endpoint"""
    print_section("2. GET SUPPORTED CLIENTS")
    try:
        resp = requests.get(f"{BASE_URL}/clients")
        print(f"Status: {resp.status_code}")
        data = resp.json()
        print(f"Response: {json.dumps(data, indent=2)}")
        if resp.status_code == 200:
            clients = data.get('data', {}).get('clients', [])
            print(f"\n✓ Found {len(clients)} supported clients:")
            for client in clients:
                print(f"  - {client['id']}: {client['name']}")
            return clients
        return []
    except Exception as e:
        print(f"❌ Error: {e}")
        return []

def test_create_session():
    """Test create upload session"""
    print_section("3. CREATE UPLOAD SESSION")
    try:
        payload = {
            "metadata": {"project": "test", "source": "test_script"},
            "ttl_hours": 24,
            "client_id": 1
        }
        resp = requests.post(
            f"{BASE_URL}/uploads/session",
            json=payload,
            headers={"Content-Type": "application/json"}
        )
        print(f"Status: {resp.status_code}")
        data = resp.json()
        print(f"Response: {json.dumps(data, indent=2)}")
        if resp.status_code == 200:
            session_id = data.get('session_id')
            print(f"✓ Session created: {session_id}")
            return session_id
        return None
    except Exception as e:
        print(f"❌ Error: {e}")
        return None

def test_get_session(session_id):
    """Test get session by ID"""
    print_section("4. GET SESSION DETAILS")
    try:
        resp = requests.get(f"{BASE_URL}/uploads/session/{session_id}")
        print(f"Status: {resp.status_code}")
        data = resp.json()
        print(f"Response: {json.dumps(data, indent=2, default=str)}")
        return resp.status_code == 200
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_upload_po_file(po_file_path, client_id=1):
    """Test PO file upload endpoint"""
    print_section(f"5. UPLOAD PO FILE (Client {client_id})")
    try:
        with open(po_file_path, 'rb') as f:
            files = {'file': (Path(po_file_path).name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            params = {
                'client_id': client_id,
                'project_id': 1,
                'uploaded_by': 'test_script',
                'auto_save': True
            }
            resp = requests.post(
                f"{BASE_URL}/po/upload",
                files=files,
                params=params
            )
        print(f"Status: {resp.status_code}")
        data = resp.json()
        print(f"Response: {json.dumps(data, indent=2, default=str)}")
        if resp.status_code == 200:
            print(f"✓ PO uploaded and parsed successfully")
            return data.get('client_po_id')
        else:
            print(f"❌ Upload failed")
            return None
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return None

def test_list_sessions():
    """Test list sessions endpoint"""
    print_section("6. LIST UPLOAD SESSIONS")
    try:
        resp = requests.get(f"{BASE_URL}/uploads/sessions")
        print(f"Status: {resp.status_code}")
        if resp.status_code == 200:
            data = resp.json()
            print(f"Response: {json.dumps(data, indent=2, default=str)}")
            return True
        else:
            # May not exist yet, just log
            print(f"Endpoint may not be implemented: {resp.status_code}")
            return False
    except Exception as e:
        print(f"Note: {e}")
        return False

def test_get_client_po(po_id):
    """Test get client PO by ID"""
    print_section("7. GET CLIENT PO DETAILS")
    try:
        resp = requests.get(f"{BASE_URL}/client-po/{po_id}")
        print(f"Status: {resp.status_code}")
        data = resp.json()
        print(f"Response: {json.dumps(data, indent=2, default=str)}")
        return resp.status_code == 200
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def main():
    """Run all tests"""
    print("\n" + "="*60)
    print("  FILE UPLOAD SYSTEM - COMPREHENSIVE TEST SUITE")
    print("="*60)
    
    # Create sample file
    po_file = create_sample_bajaj_po()
    
    try:
        # Test 1: Health check
        if not test_health_check():
            print("❌ Server not responding. Start the server first!")
            return
        
        # Test 2: Get supported clients
        clients = test_get_clients()
        
        # Test 3: Create session
        session_id = test_create_session()
        
        if session_id:
            # Test 4: Get session
            test_get_session(session_id)
        
        # Test 5: Upload PO file (most important test!)
        if os.path.exists(po_file):
            po_id = test_upload_po_file(po_file, client_id=1)
            
            # Test 6: List sessions
            test_list_sessions()
            
            # Test 7: Get uploaded PO
            if po_id:
                test_get_client_po(po_id)
        
        print_section("TEST SUMMARY")
        print("✓ All basic tests completed")
        print("\nTo test more features:")
        print("  1. Check /api/po/* endpoints for PO management")
        print("  2. Check /api/uploads/* endpoints for file management")
        print("  3. Check /api/client-po/* endpoints for client PO data")
        
    finally:
        # Cleanup
        if os.path.exists(po_file):
            os.remove(po_file)
            print(f"\n✓ Cleaned up temp file: {po_file}")

if __name__ == "__main__":
    main()
