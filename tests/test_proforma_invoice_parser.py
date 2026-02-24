# test_proforma_invoice_parser.py
"""
Test suite for Proforma Invoice Parser

Run: python -m pytest test_proforma_invoice_parser.py -v
"""

import pytest
import tempfile
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.utils.proforma_invoice_parser import (
    parse_proforma_invoice,
    ProformaInvoiceParserError,
    clean,
    to_float,
    parse_date
)


class TestUtilityFunctions:
    """Test helper functions"""
    
    def test_clean_removes_extra_spaces(self):
        assert clean("  Hello   World  ") == "Hello World"
    
    def test_clean_removes_newlines(self):
        assert clean("Hello\nWorld") == "Hello World"
    
    def test_clean_handles_none(self):
        assert clean(None) == ""
    
    def test_to_float_with_comma(self):
        assert to_float("5,486.00") == 5486.0
        assert to_float("403,040") == 403040.0
    
    def test_to_float_invalid(self):
        assert to_float("invalid") is None
    
    def test_parse_date_dd_mm_yyyy(self):
        assert parse_date("22-10-2025") == "2025-10-22"
    
    def test_parse_date_dd_slash_mm(self):
        assert parse_date("22/10/2025") == "2025-10-22"
    
    def test_parse_date_invalid(self):
        assert parse_date("invalid") is None


class TestProformaInvoiceParser:
    """Test main parser functionality"""
    
    @pytest.fixture
    def sample_pi_file(self):
        """Create a sample Proforma Invoice Excel file for testing"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Header section
        ws['A1'] = "PROFORMA INVOICE"
        ws['A3'] = "Vendor Co. Name"
        ws['B3'] = "UNV NEXGEN EXIM PVT. LTD."
        
        ws['A4'] = "Vendor Address"
        ws['B4'] = "708, 7th Floor, Palm Spring Centre, Malad West, Mumbai"
        
        ws['A5'] = "Vendor GSTIN"
        ws['B5'] = "27AABCU3488N1ZO"
        
        ws['A7'] = "Ref PO No"
        ws['B7'] = "PO01587"
        
        ws['A8'] = "PI No"
        ws['B8'] = "0036/Dava/2025-2026"
        
        ws['A9'] = "PI Date"
        ws['B9'] = "22-10-2025"
        
        ws['A12'] = "Bill To Name"
        ws['B12'] = "DAVAINDIA HEALTH MART LTD."
        
        ws['A13'] = "Bill To Address"
        ws['B13'] = "Nagpur, Maharashtra – 440012"
        
        ws['A14'] = "Bill To GSTIN"
        ws['B14'] = "27AAHCD5973D1ZI"
        
        ws['A17'] = "Ship To Name"
        ws['B17'] = "DAVAINDIA HEALTH MART LTD."
        
        ws['A18'] = "Store ID"
        ws['B18'] = "CMHNAS1747"
        
        ws['A19'] = "Site Name"
        ws['B19'] = "Takali Road, Dwarka, Nashik Maharashtra"
        
        # BOQ Header
        ws['A35'] = "Sr No"
        ws['B35'] = "BOQ Name"
        ws['C35'] = "HSN Code"
        ws['D35'] = "Qty"
        ws['E35'] = "Unit"
        ws['F35'] = "Rate"
        ws['G35'] = "Total"
        ws['H35'] = "Tax Rate"
        ws['I35'] = "Total with GST"
        
        # BOQ Line Items
        ws['A36'] = 1
        ws['B36'] = "GYPSUM FALSE CEILING"
        ws['C36'] = "68091100"
        ws['D36'] = 381
        ws['E36'] = "SQFT"
        ws['F36'] = 80
        ws['G36'] = 30480
        ws['H36'] = 5486
        ws['I36'] = 35966
        
        ws['A37'] = 2
        ws['B37'] = "INTERNAL BRANDING"
        ws['C37'] = "32091100"
        ws['D37'] = 50
        ws['E37'] = "SQFT"
        ws['F37'] = 500
        ws['G37'] = 25000
        ws['H37'] = 4500
        ws['I37'] = 29500
        
        # Tax Summary
        ws['A50'] = "Subtotal"
        ws['B50'] = 55480
        
        ws['A51'] = "CGST 9%"
        ws['B51'] = 4993.20
        
        ws['A52'] = "SGST 9%"
        ws['B52'] = 4993.20
        
        ws['A53'] = "IGST"
        ws['B53'] = 0
        
        ws['A55'] = "Grand Total"
        ws['B55'] = 65466.40
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        yield temp_file.name
        
        # Cleanup
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
    
    def test_parse_valid_proforma_invoice(self, sample_pi_file):
        """Test parsing a valid Proforma Invoice"""
        parsed = parse_proforma_invoice(sample_pi_file)
        
        # Verify structure
        assert "po_details" in parsed
        assert "line_items" in parsed
        assert "line_item_count" in parsed
        
        # Verify header fields
        po = parsed["po_details"]
        assert po["vendor_name"] == "UNV NEXGEN EXIM PVT. LTD."
        assert po["client_name"] == "DAVAINDIA HEALTH MART LTD."
        assert po["client_po_number"] == "PO01587"
        assert po["pi_number"] == "0036/Dava/2025-2026"
        assert po["pi_date"] == "2025-10-22"
        assert po["bill_to_gstin"] == "27AAHCD5973D1ZI"
        assert po["vendor_gstin"] == "27AABCU3488N1ZO"
        
        # Verify line items
        assert len(parsed["line_items"]) == 2
        assert parsed["line_item_count"] == 2
        
        item1 = parsed["line_items"][0]
        assert item1["boq_name"] == "GYPSUM FALSE CEILING"
        assert item1["hsn_code"] == "68091100"
        assert item1["quantity"] == 381
        assert item1["unit"] == "SQFT"
        assert item1["rate"] == 80
        assert item1["taxable_amount"] == 30480
        assert item1["gst_amount"] == 5486
        assert item1["gross_amount"] == 35966
        
        # Verify tax summary
        assert po["subtotal"] == 55480
        assert po["cgst"] == 4993.20
        assert po["sgst"] == 4993.20
        assert po["total_amount"] == 65466.40
    
    def test_parse_invalid_file(self):
        """Test parsing non-existent file"""
        with pytest.raises(Exception):  # Will raise FileNotFoundError or similar
            parse_proforma_invoice("/invalid/path/file.xlsx")
    
    def test_line_item_count(self, sample_pi_file):
        """Test line item counting"""
        parsed = parse_proforma_invoice(sample_pi_file)
        assert parsed["line_item_count"] == len(parsed["line_items"])


class TestValidationRules:
    """Test validation logic"""
    
    def test_validate_required_fields(self):
        """Test that validation catches missing fields"""
        invalid_parsed = {
            "po_details": {
                "vendor_name": "Test Vendor",
                # Missing client_name
                "pi_number": "PI001"
                # Missing pi_date
            },
            "line_items": [
                {
                    "boq_name": "Item 1",
                    "quantity": 10,
                    "rate": 100,
                    "gross_amount": 1000
                }
            ]
        }
        
        from app.apis.proforma_invoice import _validate_proforma_invoice
        errors = _validate_proforma_invoice(invalid_parsed)
        
        assert len(errors) > 0
        assert any("Client name" in e for e in errors)
        assert any("PI date" in e for e in errors)
    
    def test_validate_line_items(self):
        """Test line item validation"""
        invalid_parsed = {
            "po_details": {
                "vendor_name": "Test Vendor",
                "client_name": "Test Client",
                "pi_number": "PI001",
                "pi_date": "2025-01-15",
                "bill_to_gstin": "27TEST1234567890",
                "vendor_gstin": "27TEST1234567890"
            },
            "line_items": [
                {
                    "boq_name": "",  # Empty
                    "quantity": 0,    # Invalid
                    "rate": 0,        # Invalid
                    "gross_amount": 0 # Invalid
                }
            ]
        }
        
        from app.apis.proforma_invoice import _validate_proforma_invoice
        errors = _validate_proforma_invoice(invalid_parsed)
        
        assert len(errors) > 0
        assert any("BOQ name" in e for e in errors)


class TestDataMapping:
    """Test mapping to client_po schema"""
    
    def test_field_mapping(self, sample_pi_file):
        """Test that parsed fields map correctly to DB schema"""
        parsed = parse_proforma_invoice(sample_pi_file)
        po = parsed["po_details"]
        
        # These fields should be directly insertable
        assert isinstance(po.get("client_po_number"), str)
        assert isinstance(po.get("pi_number"), str)
        assert isinstance(po.get("pi_date"), str)
        assert isinstance(po.get("vendor_gstin"), str)
        assert isinstance(po.get("bill_to_gstin"), str)
        assert isinstance(po.get("subtotal"), (int, float))
        assert isinstance(po.get("cgst"), (int, float, type(None)))
    
    def test_line_item_field_mapping(self, sample_pi_file):
        """Test line item field mapping to DB schema"""
        parsed = parse_proforma_invoice(sample_pi_file)
        items = parsed["line_items"]
        
        for item in items:
            assert "boq_name" in item
            assert "hsn_code" in item
            assert "quantity" in item
            assert "unit" in item
            assert "rate" in item
            assert "taxable_amount" in item
            assert "gst_amount" in item
            assert "gross_amount" in item


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
